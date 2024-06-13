import json
import os
import re
import shutil
from datetime import datetime

import pandas
import requests
from lxml import etree
from openpyxl import load_workbook

from PC_05_91GCYY.Config import URL_HOST, DIR_M3U8, DIR_IMG, FILE_JSON_CURRENT, FILE_JSON_ALL, FILE_EXCEL_CURRENT, \
    FILE_EXCEL_ALL, DIR_OUTPUT, HEADERS


from PC_00_Common import LogUtil
from PC_00_Common.Config.StartPoint import START_POINT_1, START_POINT_2, START_POINT_3
from PC_00_Common.LogUtil.LogUtil import process_log, com_log
from PC_00_Common.ReqUtil.SavePicUtil import SavePicUtil
from PC_00_Common.XpathUtil.XpathUtil import xpath_util

# 创建session对象
from PC_00_Common.ReqUtil.ReqUtil import ReqUtil

session = requests.Session()
session.headers = HEADERS
session.cookies.set('existmag', 'all')
session.mount('https://', requests.adapters.HTTPAdapter(pool_connections=10, pool_maxsize=50))

save_pic_util = SavePicUtil(session=session)
req_util = ReqUtil(session=session)


class Task:

    def __init__(self, video_id: str, href, title, category, img_src, m3u8_url: str,
                 vip_type: str = '', play_times: str = '', data_ratio: str = '', page_order: str = '') -> None:
        super().__init__()
        self.video_id = video_id
        self.href = href
        self.title = title
        self.category = category
        self.img_src = img_src
        self.m3u8_url = m3u8_url
        self.host = URL_HOST
        self.vip_type = vip_type
        self.play_times = play_times
        self.data_ratio = data_ratio
        self.page_order = page_order

    def __str__(self):
        return json.dumps(self.__dict__, ensure_ascii=False)

    def get_name(self):
        return f'{self.video_id}_-_{self.title}_-_{self.vip_type}_-_{self.m3u8_url.split(".m3u8?val=")[1]}'


class ExecutorNoPage:
    def __init__(self):
        self.task_list: [Task] = []
        self.all_json_list = []

    def start(self):

        # 读取现有JSON文件中的数据
        if os.path.isfile(FILE_JSON_ALL):
            with open(FILE_JSON_ALL, 'r', encoding='utf-8') as json_file:
                self.all_json_list = json.load(json_file)
        category_list = [
            (0, 1, '呦呦')
        ]

        for index, category in enumerate(category_list):
            LogUtil.LogUtil.LOG_PROCESS_1 = index + 1
            LogUtil.LogUtil.LOG_PROCESS_2 = 0
            LogUtil.LogUtil.LOG_PROCESS_3 = 0
            LogUtil.LogUtil.LOG_PROCESS_4 = 0
            if index < START_POINT_1:
                process_log.process1(f"跳过 获取分类数据: {category}")
                continue
            process_log.process1(msg=f'获取分类数据 Start: {category}')
            self.read_category(category[1])
            process_log.process1(msg=f'获取分类数据 End: {category}')
        self.save_data()
        self.get_task_resource()

    def read_category(self, category):
        for video_id in range(3038, 5407):
            LogUtil.LogUtil.LOG_PROCESS_2 = video_id
            LogUtil.LogUtil.LOG_PROCESS_3 = 0
            LogUtil.LogUtil.LOG_PROCESS_4 = 0
            if video_id < START_POINT_2:
                process_log.process2(f"跳过 获取影片详情: video_id={video_id}")
                continue
            process_log.process2(f'获取影片详情 Start: video_id={video_id}')
            self.get_video_detail(video_id=f'{video_id}', category=category, page_order=f'未经过page, 未获取')
            process_log.process2(f'获取影片详情 End: video_id={video_id}')

    def get_video_detail(self, video_id, category, page_order: str = ''):
        href = f"/vid/{video_id}.html"
        res_video_page = req_util.try_get_req_times(f'{URL_HOST}{href}')
        if not res_video_page:
            com_log.error(f'获取影片详情失败: video_id={video_id}')
            return
        res_video_page_etree = etree.HTML(res_video_page.text)
        title = xpath_util.get_unique(res_video_page_etree, xpath='//div/h1[contains(@class, "mb10")]/text()',
                                      msg='获取 title')
        data_ratio = '未经过page, 未获取'
        script = xpath_util.get_unique(res_video_page_etree, xpath='//body/script/text()', msg='获取 script')
        m3u8_url = re.search(r"const m3u8_url = '(.+)';", script).group(1)
        img_src = re.search(r"const pic_thumbnail = '(.+)';", script).group(1)
        vip_type = '大会员'
        play_times = '未经过page, 未获取'

        task = Task(video_id=f'{video_id}', href=href, title=title, category=category, data_ratio=data_ratio, img_src=img_src,
                    vip_type=vip_type, play_times=play_times, m3u8_url=m3u8_url,
                    page_order=page_order)
        com_log.info(f'获取影片详情成功: video_id={video_id}, task={task}')
        self.task_list.append(task)
        return task

    def get_task_resource(self, task_list=None):
        task_list = task_list or self.task_list
        for task in task_list:
            self.get_task_m3u8(task)
            self.get_task_img(task)

    def get_task_m3u8(self, task: Task):
        res_get_m3u8 = req_util.try_get_req_times(f'{URL_HOST}{task.m3u8_url}', msg=f'获取.m3u8文件')
        if res_get_m3u8.status_code != 200:
            return
        with open(os.path.join(DIR_M3U8, f'{task.get_name()}.m3u8'), 'wb') as m3u8_file:
            m3u8_file.write(res_get_m3u8.content)

    def get_task_img(self, task: Task):
        save_pic_util.save_pic(task.img_src, DIR_IMG, task.get_name())

    # 保存到 Excel 和 Json 中
    def save_data(self, task_list=None):
        task_list = task_list or self.task_list
        # 保存 Json 文件 ↓↓↓
        # 输出本次 Json 文件 ↓↓↓
        print([task.__dict__ for task in task_list])
        print(type([task.__dict__ for task in task_list]))
        for item in [task.__dict__ for task in task_list]:
            print(type(item), item)
        with open(FILE_JSON_CURRENT, 'w', encoding='utf-8') as f:
            json.dump([task.__dict__ for task in task_list], f, ensure_ascii=False, indent=4) # ensure_ascii=False 可以保留非ASCII字符

        process_log.process3(f'输出本次Json成功: {FILE_JSON_CURRENT}')
        # 输出本次 Json 文件 ↑↑↑

        # 将本次数据合并到现有 Json 中 ↓↓↓
        # 创建一个空集合，用于存放已存在数据的序列化字符串
        existing_serialized = set()

        # 将现有数据的每个字典序列化为字符串，并加入到集合中
        for item in self.all_json_list:
            serialized_item = json.dumps(item, sort_keys=True)
            existing_serialized.add(serialized_item)

        # 遍历新数据，仅保留不在已存在数据中的字典
        unique_new_json_list = []
        for item in [task.__dict__ for task in task_list]:
            serialized_item = json.dumps(item, sort_keys=True)
            if serialized_item not in existing_serialized:
                unique_new_json_list.append(item)

        # 将新数据追加到现有数据中
        self.all_json_list = self.all_json_list + unique_new_json_list


        print(self.all_json_list)
        print(type(self.all_json_list))
        for item in self.all_json_list:
            print(type(item), item)

        # 将合并后的数据写回JSON文件
        with open(FILE_JSON_ALL, "w", encoding='utf-8') as f:
            json.dump(self.all_json_list, f, ensure_ascii=False, indent=4)  # indent 参数可选，用于美化输出（增加缩进）
        process_log.process3(f'输出所有Json成功: {FILE_JSON_ALL}')
        # 将本次数据合并到现有 Json 中 ↑↑↑
        # 保存 Json 文件 ↑↑↑

        # 保存 Excel 文件 ↓↓↓
        # 输出本次 Excel 文件 ↓↓↓
        # 将列表转换为 DataFrame , 格式为 str , 避免 Excel 以数字和科学计数法显示
        df = pandas.DataFrame(([task.__dict__ for task in task_list])).astype(str)
        # 将DataFrame保存到Excel文件
        df.to_excel(FILE_EXCEL_CURRENT, index=False)  # index=False表示不保存行索引到Excel文件
        # 加载刚才写入的Excel文件
        wb = load_workbook(FILE_EXCEL_CURRENT)
        ws = wb.active
        # 遍历所有单元格，设置格式为文本
        for row in ws.iter_rows():
            for cell in row:
                cell.number_format = '@'  # '@' 表示文本格式
        # 保存修改后的Excel文件
        wb.save(FILE_EXCEL_CURRENT)
        process_log.process3(f'输出本次Excel成功: {FILE_EXCEL_CURRENT}')
        # 输出本次 Excel 文件 ↑↑↑

        # 输出所有 Excel 文件 ↓↓↓
        df = pandas.DataFrame(self.all_json_list).astype(str)
        # 将DataFrame保存到Excel文件
        df.to_excel(FILE_EXCEL_ALL, index=False)  # index=False表示不保存行索引到Excel文件
        # 加载刚才写入的Excel文件
        wb = load_workbook(FILE_EXCEL_ALL)
        ws = wb.active
        # 遍历所有单元格，设置格式为文本
        for row in ws.iter_rows():
            for cell in row:
                cell.number_format = '@'  # '@' 表示文本格式
        # 保存修改后的Excel文件
        wb.save(FILE_EXCEL_ALL)
        process_log.process3(f'输出所有Excel成功: {FILE_EXCEL_ALL}')
        # 输出所有 Excel 文件 ↑↑↑
        # 保存 Excel 文件 ↑↑↑
        backup_folder_name = os.path.join(DIR_OUTPUT, f'Backup_{datetime.now().strftime("%Y-%m-%d_%H%M%S")}')
        os.makedirs(backup_folder_name)
        shutil.copy2(FILE_JSON_CURRENT, backup_folder_name)
        shutil.copy2(FILE_JSON_ALL, backup_folder_name)
        shutil.copy2(FILE_EXCEL_CURRENT, backup_folder_name)
        shutil.copy2(FILE_EXCEL_ALL, backup_folder_name)


if __name__ == '__main__':
    executor_no_page = ExecutorNoPage()
    executor_no_page.start()
