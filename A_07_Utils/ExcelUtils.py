from openpyxl import load_workbook


def format_excel_text(excel_path):
    # 加载刚才写入的Excel文件
    wb = load_workbook(excel_path)
    ws = wb.active
    # 遍历所有单元格，设置格式为文本
    for row in ws.iter_rows():
        for cell in row:
            cell.number_format = '@'  # '@' 表示文本格式
    # 保存修改后的Excel文件
    wb.save(excel_path)