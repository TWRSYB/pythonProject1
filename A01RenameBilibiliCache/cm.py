import json
import os
import re
import shutil
import time
from datetime import datetime

import pandas as pd

from numpy import fromfile, uint8
from openpyxl import load_workbook

from A03Common.InputCheck import input_check


def validate(file_name_str):
    r_str = r"[\/\\\:\*\?\"\<\>\|]"
    new_title = re.sub(r_str, "_", file_name_str)
    return new_title


class Target(object):

    def __init__(self, cache_type, bvid, name, video_path, audio_path, root, _json) -> None:
        super().__init__()
        self.cache_type = cache_type
        self.bvid = bvid
        self.name = name
        self.video_path = video_path
        self.audio_path = audio_path
        self.root = root
        self._json = _json

    def __str__(self) -> str:
        return f"bvid: {self.bvid}, name:{self.name}, _json: {self._json}, video_path:{self.video_path}, audio_path:{self.audio_path}, root:{self.root}"


# 执行工作类
class RenameWorker(object):

    def __init__(self, cache_path, logger) -> None:
        super().__init__()
        self.cache_path = cache_path
        self.logger = logger
        self.task_id = None
        self.task_list = []
        self.directory = None
        self.target_video_name = "video.m4s"
        self.target_audio_name = "audio.m4s"
        self.target_pc_file_info = ".videoInfo"
        self.json_list = []

    # 读取缓存方法
    def read_cache_path(self):
        self.search_path(self.cache_path)
        self.logger.info(f"识别到: {len(self.task_list)}个可合并视频。")

        # 保存到 Excel 中
        self.save_to_excel_and_json()

        for idx, task in enumerate(self.task_list):
            print(task)

        self.rename_merged_mp4()

    # 递归扫描目录下文件并匹配关键文件
    def search_path(self, root_path):
        for _l in os.listdir(root_path):
            path = os.path.join(root_path, _l)
            if os.path.isdir(path):
                self.search_path(path)
            else:
                file_info = None
                # 找到video.m4s, 调用 Android 加载文件
                if path.__contains__(self.target_video_name):
                    file_info = self.android_load_file(path)
                # 找到.videoInfo, 调用 PC 加载文件
                if path.__contains__(self.target_pc_file_info):
                    file_info = self.pc_load_file(path)
                if file_info is not None:
                    self.json_list.append(file_info)

    # Android 加载文件
    def android_load_file(self, path):
        parent_path = os.path.abspath(os.path.join(path, ".."))
        name = ""
        bvid = ""
        file_info = None
        # 读取 Json , 解析视频信息
        file_info_path = "{}/../{}".format(parent_path, "entry.json")
        try:
            with open(file_info_path, "r", encoding="utf-8") as f:
                file_info = json.load(f)
                name = self.android_parse_video_name(file_info)
                name = validate(name)
                bvid = self.android_parse_bvid(file_info)
        except Exception as e:
            self.logger.error(str(e))
        wholeness, video_path, audio_path = self.android_detect_wholeness(parent_path)
        if not wholeness:
            return

        target = Target('Android', bvid, name, video_path, audio_path, parent_path, file_info)
        self.task_list.append(target)
        return file_info

    # Android 解析名称
    def android_parse_video_name(self, file_info):
        _name = ""
        if "ep" in file_info.keys():  # 包含 ep
            self.logger.error(f"包含 ep: {file_info}")
            _name = "({}){}".format(file_info["ep"]["index"], file_info["ep"]["index_title"])
        elif "page_data" in file_info.keys():  # 包含 page_data
            if "download_subtitle" in (file_info["page_data"]).keys():
                _name = file_info["page_data"]["download_subtitle"]
                if "page" in (file_info["page_data"]).keys():
                    _page_num = file_info["page_data"]["page"]
                    _name = "{}_P{}".format(_name, _page_num)
            else:
                _name = file_info["title"]
        else:
            self.logger.error(f"既没有 ep 也没有 page_data: {file_info}")
            _name = file_info["title"]
        if len(_name) < 1:
            self.logger.error(f"无法解析名称: {file_info}")
            _name = "无法解析名称" + str(time.time())
        return _name

    # Android 获取 bvid
    def android_parse_bvid(self, file_info):
        _bvid = ""
        if "bvid" in file_info.keys():
            _bvid = file_info["bvid"]
        elif "ep" in file_info.keys() and "bvid" in (file_info["ep"]).keys():
            _bvid = file_info["ep"]["bvid"]
        if len(_bvid) < 1:
            _bvid = "无法解析bvid" + str(time.time())
        return _bvid

    # Android 检测完整性
    def android_detect_wholeness(self, parent_path):
        video_path = "{}/{}".format(parent_path, self.target_video_name)
        audio_path = "{}/{}".format(parent_path, self.target_audio_name)
        video_exist = os.path.exists(video_path)
        audio_exist = os.path.exists(audio_path)
        if video_exist is False or audio_exist is False:
            self.logger.error("FILE PATH {}:{}".format(video_exist, video_path))
            self.logger.error("FILE PATH {}:{}".format(audio_exist, audio_path))
            self.logger.info(f"missing file detected in cache path: 检测到缓存存在缺失,路径：{parent_path}")
            return False, video_path, audio_path
        return True, video_path, audio_path

    # PC 加载文件
    def pc_load_file(self, path):
        parent_path = os.path.abspath(os.path.join(path, ".."))
        file_info_path = "{}/{}".format(parent_path, self.target_pc_file_info)
        name = 'default'
        bvid = 'default'
        cid = 'default'
        try:
            with open(file_info_path, 'r', encoding='utf-8') as f:
                file_info = json.load(f)
                name = self.pc_parse_video_name(file_info)
                name = validate(name)
                bvid = self.pc_parse_bvid(file_info)
                cid = str(self.pc_parse_cid(file_info))
        except Exception as e:
            self.logger.error(str(e))
        wholeness, video_path, audio_path = self.pc_detect_loss(parent_path, cid)
        if not wholeness:
            return
        self.task_list.append(
            Target('PC', bvid, name, video_path, audio_path, parent_path, file_info)
        )

    def pc_parse_video_name(self, file_info):
        _name = ''
        if 'title' in file_info.keys():
            _name = file_info['title']
        if 'uname' in file_info.keys():
            _name = _name + '_aut{}'.format(str(file_info['uname']))
        if 'p' in file_info.keys():
            _name = _name + '_P{}'.format(str(file_info['p']))
        if len(_name) < 1:
            _name = "无法解析名称" + str(time.time())
        return _name

    def pc_parse_bvid(self, file_info):
        _bvid = ''
        if 'bvid' in file_info.keys():
            _bvid = file_info['bvid']
        if len(_bvid) < 1:
            _bvid = "无法解析bvid" + str(time.time())
        return _bvid

    def pc_detect_loss(self, _parent_path, _cid):
        cache_file = []
        video_path = ''
        audio_path = ''
        for _l in os.listdir(_parent_path):
            path = os.path.join(_parent_path, _l)
            if os.path.isdir(path):
                continue
            if path.__contains__(_cid) and path.__contains__('.m4s'):
                cache_file.append(path)
        if len(cache_file) != 2:
            self.logger.error('pc cache elems error :{},{}'.format(cache_file, str(cache_file)))
            return False, video_path, audio_path
        if os.stat(cache_file[0]).st_size > os.stat(cache_file[1]).st_size:
            return True, cache_file[0], cache_file[1]
        return True, cache_file[1], cache_file[0]

    def pc_parse_cid(self, file_info):
        if 'cid' in file_info.keys():
            return file_info['cid']
        if 'itemId' in file_info.keys():
            return file_info['itemId']

    def pc_cache_decode(self, _path_input, _path_output):
        read = fromfile(_path_input, dtype=uint8)
        if all(read[0:9] == [48, 48, 48, 48, 48, 48, 48, 48, 48]):
            read[9:].tofile(_path_output)
        elif all(read[0:3] == [255, 255, 255]):
            read[3:].tofile(_path_output)
        else:
            self.logger.error('maybe has new change: {}'.format(str(read[0:64])))

    # 保存到 Excel
    def save_to_excel_and_json(self):
        # 获取当前日期和时间
        backup_folder_name = f'./OutputData/Backup_{datetime.now().strftime("%Y-%m-%d_%H%M%S")}'
        current_excel_name = f'./OutputData/Current_Excel.xlsx'
        current_json_name = f'./OutputData/Current_JSON.json'
        all_excel_name = f'./OutputData/All_Excel.xlsx'
        all_json_name = f'./OutputData/All_JSON.json'

        # 保存 Json 文件 ↓↓↓
        # 输出本次 Json 文件 ↓↓↓
        with open(current_json_name, 'w', encoding='utf-8') as f:
            json.dump(self.json_list, f, ensure_ascii=False, indent=4)
        self.logger.info(f'输出本次Json成功: {current_json_name}')
        # 输出本次 Json 文件 ↑↑↑

        # 将本次数据合并到现有 Json 中 ↓↓↓
        # 读取现有JSON文件中的数据
        with open(all_json_name, 'r', encoding='utf-8') as json_file:
            existing_json_list = json.load(json_file)

        # 创建一个空集合，用于存放已存在数据的序列化字符串
        existing_serialized = set()

        # 将现有数据的每个字典序列化为字符串，并加入到集合中
        for item in existing_json_list:
            serialized_item = json.dumps(item, sort_keys=True)
            existing_serialized.add(serialized_item)

        # 遍历新数据，仅保留不在已存在数据中的字典
        unique_new_json_list = []
        for item in self.json_list:
            serialized_item = json.dumps(item, sort_keys=True)
            if serialized_item not in existing_serialized:
                unique_new_json_list.append(item)

        # 将新数据追加到现有数据中
        merged_json_list = existing_json_list + unique_new_json_list

        # 将合并后的数据写回JSON文件
        with open(all_json_name, "w") as json_file:
            json.dump(merged_json_list, json_file, indent=4)  # indent 参数可选，用于美化输出（增加缩进）
        self.logger.info(f'输出所有Json成功: {all_json_name}')
        # 将本次数据合并到现有 Json 中 ↑↑↑
        # 保存 Json 文件 ↑↑↑

        # 保存 Excel 文件 ↓↓↓
        # 输出本次 Excel 文件 ↓↓↓
        # 将列表转换为 DataFrame , 格式为 str , 避免 Excel 以数字和科学计数法显示
        df = pd.DataFrame(self.json_list).astype(str)
        # 将DataFrame保存到Excel文件
        df.to_excel(current_excel_name, index=False)  # index=False表示不保存行索引到Excel文件
        # 加载刚才写入的Excel文件
        wb = load_workbook(current_excel_name)
        ws = wb.active
        # 遍历所有单元格，设置格式为文本
        for row in ws.iter_rows():
            for cell in row:
                cell.number_format = '@'  # '@' 表示文本格式
        # 保存修改后的Excel文件
        wb.save(current_excel_name)
        self.logger.info(f'输出本次Excel成功: {current_excel_name}')
        # 输出本次 Excel 文件 ↑↑↑

        # 输出所有 Excel 文件 ↓↓↓
        df = pd.DataFrame(merged_json_list).astype(str)
        # 将DataFrame保存到Excel文件
        df.to_excel(all_excel_name, index=False)  # index=False表示不保存行索引到Excel文件
        # 加载刚才写入的Excel文件
        wb = load_workbook(all_excel_name)
        ws = wb.active
        # 遍历所有单元格，设置格式为文本
        for row in ws.iter_rows():
            for cell in row:
                cell.number_format = '@'  # '@' 表示文本格式
        # 保存修改后的Excel文件
        wb.save(all_excel_name)
        self.logger.info(f'输出所有Excel成功: {all_excel_name}')
        # 输出所有 Excel 文件 ↑↑↑
        # 保存 Excel 文件 ↑↑↑

        os.makedirs(backup_folder_name)
        backup_current_json_path = shutil.copy2(current_json_name, backup_folder_name)
        self.logger.info(f'备份本次Json成功: {backup_current_json_path}')
        backup_all_json_path = shutil.copy2(all_json_name, backup_folder_name)
        self.logger.info(f'备份合并Json成功: {backup_all_json_path}')
        backup_current_excel_path = shutil.copy2(current_excel_name, backup_folder_name)
        self.logger.info(f'备份本次Excel成功: {backup_current_excel_path}')
        backup_all_excel_path = shutil.copy2(all_excel_name, backup_folder_name)
        self.logger.info(f'备份所有Excel成功: {backup_all_excel_path}')

    def rename_merged_mp4(self):
        mp4_dir = input_check(msg="如需重命名已合并的视频, 请输入合并视频的目录, 否则输入N, 请输入:", not_null=True)
        if mp4_dir == 'N':
            return
        if os.path.isdir(mp4_dir):
            pass
